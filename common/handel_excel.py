#!/usr/bin/env python
# -*- coding:utf-8 -*-
'''
#@Time      :2022/4/24 22:21
#@Author    :xiaoxin
#@File      :handel_excel.py
'''
import openpyxl

class HandleExcel:

    def __init__(self,filepath,sheetname):
        """
              :param filepath: elcle文件路径
              :param sheetname: sheet表单名
              """
        self.filepath = filepath
        self.sheetname = sheetname

    def read_excel(self):
        """读取excel数据"""
        #加载excel文件
        workbook = openpyxl.load_workbook(self.filepath)
        sheet = workbook[self.sheetname]
        sheet_rows = list(sheet.rows)
        # 获取第一行作为标题
        title = [i.value for i in sheet_rows[0]]
        cases = []
        # 获取第一行外的所有行
        for item in sheet_rows[1:]:
            data = [i.value for i in item]
            case = dict(zip(title,data))
            cases.append(case)
        return  cases

    def write_excel(self,row, column, value):
        workbook = openpyxl.load_workbook(self.filepath)
        sheet = workbook[self.sheetname]
        #写入数据保存文件
        sheet.cell(row=row, column=column, value=value)
        workbook.save(self.filepath)